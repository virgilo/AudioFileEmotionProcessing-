#Fonctions Utiles
##Excel Debut
#Required Librairies
import os
from openpyxl import Workbook,load_workbook
#Creer un nouveau Classeur
def CreerNewClasseur(NomDuFicherExcel):
    ConteneursDesClasseurs = Workbook()
    ConteneursDesClasseurs.save(NomDuFicherExcel)
#Ecrire une liste/array dans python sur la ligne
def HorizontalWritingXlsx(NomDuFicherExcel,NomDeLaFeuilDeCalcul,CelluleDeDepart,tableauDeDonnee):
    ClasseurCourant = load_workbook(NomDuFicherExcel)
    FeuilleCourante = ClasseurCourant[NomDeLaFeuilDeCalcul]
    indiceRow=CelluleDeDepart[0]
    indiceColumn=CelluleDeDepart[1]
    for i in range(0,len(tableauDeDonnee)):
        FeuilleCourante.cell(row=indiceRow, column=indiceColumn).value = tableauDeDonnee[i]
        indiceColumn=indiceColumn+1
    ##Enrégistrement
    ClasseurCourant.save(NomDuFicherExcel)
    

#Ecrire une liste/array dans python sur la colonne
def VerticalWritingXlsx(NomDuFicherExcel,NomDeLaFeuilDeCalcul,CelluleDeDepart,tableauDeDonnee):
    ClasseurCourant = load_workbook(NomDuFicherExcel)
    FeuilleCourante = ClasseurCourant[NomDeLaFeuilDeCalcul]
    indiceRow=CelluleDeDepart[0]
    indiceColumn=CelluleDeDepart[1]
    for i in range(0,len(tableauDeDonnee)):
        FeuilleCourante.cell(row=indiceRow, column=indiceColumn).value = tableauDeDonnee[i]
        indiceRow=indiceRow+1
    ##Enrégistrement
    ClasseurCourant.save(NomDuFicherExcel)
    
## Excel Fin



#Prerequisies
import os, sys
from pydub import AudioSegment
import numpy as np

##Fonction Conversion Audio en Array
def AudioSegToArray(Var_AudioSegment):
    audio_ARRAY=Var_AudioSegment.get_array_of_samples()
    audio_NUMPY_ARRAY=np.array(audio_ARRAY)
    return audio_ARRAY

##Fonction Conversion Audio en NumpyArray
def AudioSegToNumpyArray(Var_AudioSegment):
    audio_ARRAY=Var_AudioSegment.get_array_of_samples()
    audio_NUMPY_ARRAY=np.array(audio_ARRAY)
    audio_NUMPY_ARRAY=np.int16(audio_NUMPY_ARRAY)
    return audio_NUMPY_ARRAY

##Fonction Conversion Array en NumpyArray
def ArrayToNumpyArray(audio_ARRAY):
    audio_NUMPY_ARRAY=np.array(audio_ARRAY)
    audio_NUMPY_ARRAY=np.int16(audio_NUMPY_ARRAY)
    return audio_NUMPY_ARRAY

##Fonction Conversion NumpyArray en AudioSeg
def NumpyArrayToAudioSeg(var_NumpyArray):
    return AudioSegment(var_NumpyArray.tobytes(),frame_rate=24414,sample_width=2,channels=1)

##Fonction Conversion AudioSeg to AudioFile
def AudioSegToAudioFile(var_AudioSeg,PathFilename,i,extensionFilename):
    #("./AUDIOsignalProcessingOUPUT/AudiosMangas_{}.wav".format(i), format="wav")
    return var_AudioSeg.export(PathFilename.format(i), format=extensionFilename)

##List of wav file in a folder
def ListOfWavFileInaFolder(Folderpath):   
    ListOfWavFiles=[]
    for subdir, dirs, files in os.walk(Folderpath):
        for file in files:
            #print os.path.join(subdir, file)
            filepath = subdir + os.sep + file
            if filepath.endswith((".wav")):  #Afin d'exclure le fichier desktop.ini:  ./TESS-AUDIOS-LOCUTEUR-Y\Angry\desktop.ini
                #print(filepath)
                ListOfWavFiles.append(filepath)
    return ListOfWavFiles

##Selection of Intial utterance of the both actor that will be use to build up the artificial conversation
def Selected_Initial_Utterances_BothActors_Fct(pathActorO,pathActorY):
   
    rootdirO=pathActorO
    rootdirY=pathActorY
    ListFilesActorO=[]
    ListFilesActorY=[]

    for subdir, dirs, files in os.walk(rootdirO):
        for file in files:
            #print os.path.join(subdir, file)
            filepath = subdir + os.sep + file
            if filepath.endswith((".wav")):  #Afin d'exclure le fichier desktop.ini:  ./TESS-AUDIOS-LOCUTEUR-Y\Angry\desktop.ini
                #print(filepath)
                ListFilesActorO.append(filepath)
    for subdir, dirs, files in os.walk(rootdirY):
        for file in files:
            #print os.path.join(subdir, file)
            filepath = subdir + os.sep + file
            if filepath.endswith((".wav")):  #Afin d'exclure le fichier desktop.ini:  ./TESS-AUDIOS-LOCUTEUR-Y\Angry\desktop.ini
                #print(filepath)
                ListFilesActorY.append(filepath)

    ListFilesActorO_1stUtt=list(filter(lambda Fichier: Fichier.endswith(("neutral.wav"))==True, ListFilesActorO))
    ListFilesActorO_2ndUtt=list(filter(lambda Fichier: Fichier.endswith(("disgust.wav"))==True, ListFilesActorO))
    ListFilesActorO_3rdUtt=list(filter(lambda Fichier: Fichier.endswith(("angry.wav"))==True, ListFilesActorO))
    ListFilesActorO_4thUtt=list(filter(lambda Fichier: Fichier.endswith(("happy.wav"))==True, ListFilesActorO))
    
    ListFilesActorY_1stUtt=list(filter(lambda Fichier: Fichier.endswith(("neutral.wav"))==True, ListFilesActorY))
    ListFilesActorY_2ndUtt=list(filter(lambda Fichier: Fichier.endswith(("ps.wav"))==True, ListFilesActorY))
    ListFilesActorY_3rdUtt=list(filter(lambda Fichier: Fichier.endswith(("sad.wav"))==True, ListFilesActorY))
    ListFilesActorY_4thUtt=list(filter(lambda Fichier: Fichier.endswith(("disgust.wav"))==True, ListFilesActorY))
    
    Selected_Initial_Utterances_ActorO_Paths=[]
    Selected_Initial_Utterances_ActorO_Paths.append(ListFilesActorO_1stUtt)
    Selected_Initial_Utterances_ActorO_Paths.append(ListFilesActorO_2ndUtt)
    Selected_Initial_Utterances_ActorO_Paths.append(ListFilesActorO_3rdUtt)
    Selected_Initial_Utterances_ActorO_Paths.append(ListFilesActorO_4thUtt)
    
    Selected_Initial_Utterances_ActorY_Paths=[]
    Selected_Initial_Utterances_ActorY_Paths.append(ListFilesActorY_1stUtt)
    Selected_Initial_Utterances_ActorY_Paths.append(ListFilesActorY_2ndUtt)
    Selected_Initial_Utterances_ActorY_Paths.append(ListFilesActorY_3rdUtt)
    Selected_Initial_Utterances_ActorY_Paths.append(ListFilesActorY_4thUtt)
        
    Selected_Initial_Utterances_BothActors_Paths=[]
    Selected_Initial_Utterances_BothActors_Paths.append(Selected_Initial_Utterances_ActorO_Paths)
    Selected_Initial_Utterances_BothActors_Paths.append(Selected_Initial_Utterances_ActorY_Paths)
    
    return Selected_Initial_Utterances_BothActors_Paths
    
    

def ListInitialUttByActorandConv_Paths(IndiceActor,IndiceConversation,pathActorO,pathActorY):
    #IndiceActor=0 ou 1 pour l'acteur 1O ou 2Y
    Selected_Initial_Utterances_BothActors_Paths=Selected_Initial_Utterances_BothActors_Fct(pathActorO,pathActorY)
    Selected_Initial_Utterances_Actor_Paths=Selected_Initial_Utterances_BothActors_Paths[IndiceActor]
    resultat=[]
    for i, yy in enumerate(Selected_Initial_Utterances_Actor_Paths):
        resultat.append(Selected_Initial_Utterances_Actor_Paths[i][IndiceConversation])
    return resultat
def ListExtractedFilesByActor(PathExtractedFilesActorO):
    resultat=[]
    for subdir, dirs, files in os.walk(PathExtractedFilesActorO):
        for file in files:
            #print os.path.join(subdir, file)
            filepath = subdir + os.sep + file
            if filepath.endswith((".wav")):  #Afin d'exclure le fichier desktop.ini:  ./TESS-AUDIOS-LOCUTEUR-Y\Angry\desktop.ini
                #print(filepath)
                resultat.append(filepath)
    return resultat

def testAff():
    print('ViDsible')

    
##Trier une liste de chaine alphanumérique selon l'ordre croissant des nombres suffixes _XX    
##REMARQUE les conversations sont suffixées par des nombres entiers via le "_"

def helper_Sortfct(elt):
    NbrePointExtension=elt.split('_')[1] # exple:   0.wav , 1.wav, ... 199.wav
    Nbre , extension=NbrePointExtension.split('.')
    return int(Nbre)

def SortAlphanumericalList(L):
    L=sorted(L,reverse=False, key= lambda elt: helper_Sortfct(elt) )
    return L

##Variables

PathFilename="./ArtificialdatasetAudiosConversation/bCreatedConversation_{}.wav"
PathExtractedFilesActorO="./AudiosApresSegmentation/AudiosDuLocuteurs1"
PathExtractedFilesActorY="./AudiosApresSegmentation/AudiosDuLocuteurs2"
ExtensionFilename="wav"

##Le separateur/Marqueur - AudioSegment
Marker_AudioSeg=AudioSegment.silent(duration=2000)
#Marker_AudioSeg=AudioSegment.from_wav("./AUDIO MARKER/beepAudio.wav")

## separateur/Marqueur - Array
Marker_Array=AudioSegToArray(Marker_AudioSeg)

## Chemin des dossiers contenant les prises de parole du locuteur 1 et du locuteur 2
pathActorO = "./TESS-AUDIOS-LOCUTEUR-O"
pathActorY = "./TESS-AUDIOS-LOCUTEUR-Y"


##
Selected_Initial_Utterances_BothActors_Paths=Selected_Initial_Utterances_BothActors_Fct(pathActorO,pathActorY)
Selected_Initial_Utterances_ActorO_Paths=Selected_Initial_Utterances_BothActors_Paths[0]
Selected_Initial_Utterances_ActorY_Paths=Selected_Initial_Utterances_BothActors_Paths[1]
##Nobre de prise de parole par locuteur dans la conversation
NbreDePriseDeParoleByLoc=len(Selected_Initial_Utterances_ActorO_Paths)

#-----------------------------------------------------LoadingMyPretrainedAIModel--------------------------------------------------------------------#
import tensorflow as tf
#LoadingSomeKereasModel
def LoadingMyPretrainedAIModel():
    Pretrained_Model_PATH='C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/PretrainedCnnModel/model_4.h5'
    MonModel = tf.keras.models.load_model('MyPretrainedModel')
    return MonModel
#-----------------------------------------------------EmotionsPredictionFormSpectrogramme--------------------------------------------------------------------#
#Fonction qui prend en entrée un spectrogramme et donne en sortie une émotion
import tensorflow as tf
from tensorflow.keras.preprocessing.image import ImageDataGenerator
#Dictionnaires d'émotions
#Emotions_Dictionary={0:'Colère',1:'Crainte',2:'Dégoût',3:'Happy',4:'Neutre',5:'Surprise',6:'Triste'}
Emotions_Dictionary={0:'Neutral',1:'Happy',2:'Sad',3:'Angry',4:'Fearful',5:'Disgust',6:'pSurpise'}
def EmotionsPredictionFormSpectrogramme(SpectrogramUtteranceFolder):
    #Charger les SpectroGramme des prises de paroles (classés par ordre chrnologique_alphanumérique dans le repertoire)
    path_dir = SpectrogramUtteranceFolder
    PrisesDeParole_datagen = ImageDataGenerator(
                    rescale=1./255)
    datagen = ImageDataGenerator(
                    rescale=1./255,
                    validation_split = 0.3)
    PrisesDeParole_generator = datagen.flow_from_directory(
    path_dir,
    target_size=(350,350),
    shuffle=False)
    
    #charger le model pré entrainé
    Pretrained_Model_PATH='C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/PretrainedCnnModel/model_4.h5'
    MyPretrained_Model = tf.keras.models.load_model(Pretrained_Model_PATH)
    
    #Prediction des emotions
    pred = MyPretrained_Model.predict(PrisesDeParole_generator)
    y_pred = np.argmax(pred, axis=1)
    #Display result
    print(y_pred)
    return y_pred

#-----------------------------------------------------FormAudiosToSpectrogrammes--------------------------------------------------------------------#

#Cette fonction convertie tous les fichiers audios contenus dans un repertoire en spectrogramme dans un autre repertoire
FloderDestinationSpectrogramme='C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/FloderDestinationSpectrogramme/'
def OLD_FormAudiosToSpectrogrammes(FolderSource):
    #La transformation en spectrogramme
    from scipy.io import wavfile # scipy library to read wav files
    import numpy as np
    import os
    # Plot the audio signal in time
    import matplotlib.pyplot as plt
    
    path = os. getcwd()
    paths = FolderSource
    #paths = FolderSource+"*.wav"
    print(paths)
    dirs = os.listdir(paths)

    for i, file in enumerate(dirs):
    #AudioName = "chunk0.vaw" # Audio File
        print(paths + file)
        fs, Audiodata = wavfile.read(paths + file)

    #Spectrogram
        from scipy import signal
        N = 512 #Number of point in the fft 
        f, t, Sxx = signal.spectrogram(Audiodata, fs,window = signal.blackman(N),nfft=N)
        plt.figure()
        plt.pcolormesh(t, f,10*np.log10(Sxx)) # dB spectrogram
    #plt.pcolormesh(t, f,Sxx) # Lineal spectrogram
        plt.ylabel('Frequency [Hz]')
        plt.xlabel('Time [seg]')
        plt.title('Spectrogram with scipy.signal',size=16);
        plt.colorbar()
        plt.savefig(FloderDestinationSpectrogramme+'MySpectrogramme'+str(i)+'.png' , dpi = 72 )
        #plt.savefig.export("./colspect/Spectogramme_{i}.png' , dpi = 72 ")
        plt.show()

#-----------------------------------------------------FormAudiosToSpectrogrammes-----With Folder Destination---------------------------------------------------------------#        
def Old_FormAudiosToSpectrogrammesWithFolderDest(FolderSource,FolderDestination):
    #La transformation en spectrogramme
    from scipy.io import wavfile # scipy library to read wav files
    import numpy as np
    import os
    # Plot the audio signal in time
    import matplotlib.pyplot as plt
    
    path = os. getcwd()
    paths = FolderSource
    #paths = FolderSource+"*.wav"
    print(paths)
    dirs = os.listdir(paths)

    for i, file in enumerate(dirs):
    #AudioName = "chunk0.vaw" # Audio File
        print(paths + file)
        fs, Audiodata = wavfile.read(paths + file)

    #Spectrogram
        from scipy import signal
        N = 512 #Number of point in the fft 
        f, t, Sxx = signal.spectrogram(Audiodata, fs,window = signal.blackman(N),nfft=N)
        plt.figure()
        plt.pcolormesh(t, f,10*np.log10(Sxx)) # dB spectrogram
    #plt.pcolormesh(t, f,Sxx) # Lineal spectrogram
        plt.ylabel('Frequency [Hz]')
        plt.xlabel('Time [seg]')
        plt.title('Spectrogram with scipy.signal',size=16);
        plt.colorbar()
        plt.savefig(FolderDestination+'MySpectrogramme'+str(i)+'.png' , dpi = 72 )
        #plt.savefig.export("./colspect/Spectogramme_{i}.png' , dpi = 72 ")
        plt.show()

#-----------------------------------------------------Code to Add the folder of this module in the system path--------------------------------------------------------------------#
#Par conséquent, je pourrai importer ce module de n'importe quel repertoire
import sys
sys.path.append('C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/2-Samedi 14 Mai 2022/')
        
#--------------------------------------ALL - PATHS  SHOULD BE THERE AT THE END--------------FormAudiosToSpectrogrammes--------------------------------------------------------------------#
RavdessByEmotion_PATH="C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/DATABASEInThe_STATEofTheART/RAVDESS_By_Emotion/"
RavdessBySpectrogram_PATH="C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/DATABASEInThe_STATEofTheART/RAVDESS_by_Spectro/"

TessByEmotion_PATH="C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/DATABASEInThe_STATEofTheART/TESS_by_Emotion/"
TessBySpectrogram_PATH="C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/DATABASEInThe_STATEofTheART/TESS_by_Spectro/"



#-----------------------------------------------------IUGET___Code to Add the folder of this module in the system path--------------------------------------------------------------------#
#Par conséquent, je pourrai importer ce module de n'importe quel repertoire
import sys
sys.path.append('C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/2-Samedi 14 Mai 2022/')
        
#--------------------------------------IUGET___ALL  - PATHS  SHOULD BE THERE AT THE END--------------FormAudiosToSpectrogrammes--------------------------------------------------------------------#
RavdessByEmotion_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/RAVDESS_By_Emotion/"
RavdessByMFCCgram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/RAVDESS_by_MFCC/"
RavdessByChromagram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/RAVDESS_by_CHROMA/"

TessByEmotion_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/TESS_by_Emotion/"
TessByMFCCgram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/TESS_by_MFCC/"
TessByChromagram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/TESS_by_CHROMA/"
TessBySpectrogram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/TESS_by_Spectro/"

EmoDBByEmotion_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/EmoDB_By_Emotion/"
EmoDBByMFCCgram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/EmoDB_by_MFCC/"
EmoDBByChromagram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/EmoDB_by_CHROMA/"
EmoDBBySpectrogram_PATH_IUGET="C:/Users/INSIDE BINARIES/Documents/LABO_PYTHON/DATASET/EmoDB_By_Spectro/"

#-----------------------------------------------------IUGET___FormAudiosToASingleMFCC--------------------------------------------------------------------#
import librosa
import librosa.display
import matplotlib.pyplot as plt 
import numpy as np
## FCT To convert convert an audio to a saved spectrogram 
import sys
#sys.path.append('C:/Users/fijitsu/anaconda/envs/opensoundscape/Lib/site-packages/')
print("IUGET TRAINING_")
#print('ddd')
from pathlib import Path
#a single line to convert an audio to a saved MFCC 
#audio_filename="C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/DATABASEInThe_STATEofTheART/RAVDESS_By_Emotion/02-Calm/03-01-02-01-01-01-01.wav"
def SingleAudioToSingleMFCCgram(audio_filename,image_save_path):
    x, sr = librosa.load(audio_filename)
    mfccs = librosa.feature.mfcc(x, sr=sr)
    # Displaying  the MFCCs:
    plt.figure(figsize=(15, 3))
    #librosa.display.specshow(mfccs, sr=sr, x_axis='time')
    librosa.display.specshow(mfccs, sr=sr)
    #print(Xdb)
    plt.savefig(image_save_path , dpi = 72 )
    return mfccs

#-----------------------------------------------------IUGET___FormAudiosToMFCC-----With Folder Destination-- NEW-------------------------------------------------------------#        
def FormAudiosToMFCCgramWithFolderDest(FolderSource,FolderDestination):
    #La transformation en MFFCCgram
    from scipy.io import wavfile # scipy library to read wav files
    import numpy as np
    import os
    # Plot the audio signal in time
    import matplotlib.pyplot as plt
    
    path = os. getcwd()
    paths = FolderSource
    #paths = FolderSource+"*.wav"
    print(paths)
    dirs = os.listdir(paths)

    for i, file in enumerate(dirs):
    #AudioName = "chunk0.vaw" # Audio File
        print(paths + file)
    #Spectrogram
        audio_filename=paths + file
        if (audio_filename.endswith((".ini"))==False ): #S'assurer que le fichier traité est un
            #fichier audio, d autre extension sont à envisager
            image_save_path=FolderDestination+'MFCCgram'+str(i)+'.png' 
            print(image_save_path)
            SingleAudioToSingleMFCCgram(audio_filename,image_save_path)
            #plt.savefig.export("./colspect/Spectogramme_{i}.png' , dpi = 72 ")
#-----------------------------------------------------END------------------------------------------------------------------------------------------------# 


#-----------------------------------------------------ICI IUGET___ChromaGRAM--------------------------------------------------------------------#


#-----------------------------------------------------IUGET___FormAudiosToASingleChroma--------------------------------------------------------------------#
import librosa
import librosa.display
import matplotlib.pyplot as plt 
import numpy as np
## FCT To convert convert an audio to a saved spectrogram 
import sys
#sys.path.append('C:/Users/fijitsu/anaconda/envs/opensoundscape/Lib/site-packages/')
print("IUGET TRAINING_Chroma")
#print('ddd')
from pathlib import Path
#a single line to convert an audio to a saved MFCC 
#audio_filename="C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/DATABASEInThe_STATEofTheART/RAVDESS_By_Emotion/02-Calm/03-01-02-01-01-01-01.wav"
def SingleAudioToSingleChromagram(audio_filename,image_save_path):
    x, sr = librosa.load(audio_filename)
    hop_length = 512
    chromagram = librosa.feature.chroma_stft(x, sr=sr, hop_length=hop_length)
    fig, ax = plt.subplots(figsize=(15, 3))
    img = librosa.display.specshow(chromagram, hop_length=hop_length, cmap='coolwarm')
    plt.savefig(image_save_path , dpi = 72 )
    #fig.colorbar(img, ax=ax)
    return chromagram

#-----------------------------------------------------IUGET___FormAudiosToChroma-----With Folder Destination-- NEW-------------------------------------------------------------#        
def FormAudiosToChromagramWithFolderDest(FolderSource,FolderDestination):
    #La transformation en MFFCCgram
    from scipy.io import wavfile # scipy library to read wav files
    import numpy as np
    import os
    # Plot the audio signal in time
    import matplotlib.pyplot as plt
    
    path = os. getcwd()
    paths = FolderSource
    #paths = FolderSource+"*.wav"
    print(paths)
    dirs = os.listdir(paths)

    for i, file in enumerate(dirs):
    #AudioName = "chunk0.vaw" # Audio File
        print(paths + file)
    #Spectrogram
        audio_filename=paths + file
        if (audio_filename.endswith((".ini"))==False ): #S'assurer que le fichier traité est un
            #fichier audio, d autre extension sont à envisager
            image_save_path=FolderDestination+'Chromagram'+str(i)+'.png' 
            print(image_save_path)
            SingleAudioToSingleChromagram(audio_filename,image_save_path)
            #plt.savefig.export("./colspect/Spectogramme_{i}.png' , dpi = 72 ")
#-----------------------------------------------------END------------------------------------------------------------------------------------------------# 



#-----------------------------------------------------BEGIN-------FormAudiosToSPECTROGRAM-----With Folder Destination-----------------------------------------------------------------------______RE-COPIÉ A IUGET____MAIS A SUPPRIMÉ A LA MAISON CAR Y FIGURE DEJA-------------------------------------------------------------#  
import librosa
import librosa.display
import matplotlib.pyplot as plt 
import numpy as np
#1080 216

def SingleAudioToSingleSpectrogram(audio_data,image_save_path):
    x, sr = librosa.load(audio_data)
    # Spectrogram of frequency
    X = librosa.stft(x)
    Xdb = librosa.amplitude_to_db(abs(X))
    plt.figure(figsize=(15, 3))
    #librosa.display.specshow(Xdb, sr=sr, x_axis='time', y_axis='hz')
    #plt.colorbar()
    librosa.display.specshow(Xdb, sr=sr)
    #print(Xdb)
    plt.savefig(image_save_path, dpi = 72 )
    return Xdb


#-----------------------------------------------------FormAudiosToSpectrogrammes--------------------------------------------------------------------#

#Cette fonction convertie tous les fichiers audios contenus dans un repertoire en spectrogramme dans un autre repertoire
FloderDestinationSpectrogramme='C:/Users/fijitsu/Documents/A-LABORATORY-PYTHON/FloderDestinationSpectrogramme/'
def FormAudiosToSpectrogrammes(FolderSource):
    #La transformation en spectrogramme
    from scipy.io import wavfile # scipy library to read wav files
    import numpy as np
    import os
    # Plot the audio signal in time
    import matplotlib.pyplot as plt
    
    path = os. getcwd()
    paths = FolderSource
    #paths = FolderSource+"*.wav"
    print(paths)
    dirs = os.listdir(paths)

    for i, file in enumerate(dirs):
    #AudioName = "chunk0.vaw" # Audio File
        print(paths + file)
        #fs, Audiodata = wavfile.read(paths + file)
    #Spectrogram
        audio_filename=paths + file
        #image_save_path=FloderDestinationSpectrogramme+'OpenSpectrogram'+str(i)+'.png'  #Opensoundscape
        image_save_path=FloderDestinationSpectrogramme+'LibroSpectrogram'+str(i)+'.png'  #Librosa
        SingleAudioToSingleSpectrogram(audio_filename,image_save_path)
        #plt.savefig.export("./colspect/Spectogramme_{i}.png' , dpi = 72 ")

#-----------------------------------------------------FormAudiosToSpectrogrammes-----With Folder Destination-- NEW-------------------------------------------------------------#        
def FormAudiosToSpectrogrammesWithFolderDest(FolderSource,FolderDestination):
    #La transformation en spectrogramme
    from scipy.io import wavfile # scipy library to read wav files
    import numpy as np
    import os
    # Plot the audio signal in time
    import matplotlib.pyplot as plt
    
    path = os. getcwd()
    paths = FolderSource
    #paths = FolderSource+"*.wav"
    print(paths)
    dirs = os.listdir(paths)

    for i, file in enumerate(dirs):
    #AudioName = "chunk0.vaw" # Audio File
        print(paths + file)
    #Spectrogram
        audio_filename=paths + file
        if (audio_filename.endswith((".ini"))==False ): #S'assurer que le fichier traité est un
            #fichier audio, d autre extension sont à envisager
            #image_save_path=FolderDestination+'OpenSpectrogram'+str(i)+'.png'  #Opensoundscape
            image_save_path=FolderDestination+'LibroSpectrogram'+str(i)+'.png'  #Librosa
            print(image_save_path)
            SingleAudioToSingleSpectrogram(audio_filename,image_save_path)
            #plt.savefig.export("./colspect/Spectogramme_{i}.png' , dpi = 72 ")
#-----------------------------------------------------END------------------------------------------------------------------------------------------------# 
#-----------------------------------------------------END-------FormAudiosToSPECTROGRAM-----With Folder Destination-- ______RE-COPIÉ A IUGET _MAIS A SUPPRIMÉ A LA MAISON CAR Y FIGURE DEJA-------------------------------------------------------------#  

